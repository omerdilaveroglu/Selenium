from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from webdriver_manager.chrome import ChromeDriverManager 
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import date
import openpyxl
from constants import globalConstants as gc

class Test_DemoClass:
    #Her testten önce çağırılır
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(gc.URL)
        #Günün tarihini al bu tarih ile bir klasör var mı kontrol et yoksa oluştur.
        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)
        
    #Her testten sonra çağırılır.
    def teardown_method(self):
        self.driver.quit()
        print("")
    
    def getData():
        #veriyi al
        excelFile = openpyxl.load_workbook("data/invalid_login.xlsx")
        selectedSheet = excelFile.worksheets[0]

        totalRows = selectedSheet.max_row
        data=[]
        for i in range(2, totalRows+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)

        return data
        #return [("1","1"),("kullaniciadim","sifrem"),("Kodalamaio","123")]    

    #@pytest.mark.skip()#Bu testi atla, yapma anlamına gelir.
    @pytest.mark.parametrize("username,password",getData())
    def test_invalid_login(self,username,password):
        self.waitForElementVisible((By.ID,"user-name"),15)
        userNameInput = self.driver.find_element(By.ID,"user-name")
        self.waitForElementVisible((By.ID,"password"))    
        passwordInput = self.driver.find_element(By.ID,"password")
        userNameInput.send_keys(username)
        passwordInput.send_keys(password)
        
        self.waitForElementVisible((By.ID,"login-button"))

        loginBtn = self.driver.find_element(By.ID,"login-button").click()
        self.waitForElementVisible((By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3"))
        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png")
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"

    def waitForElementVisible(self,locator,timeout=15):
        WebDriverWait(self.driver,timeout).until(ec.visibility_of_all_elements_located(locator))

